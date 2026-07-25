"""
kb_to_servicenow.py
Converts REDCap KB markdown articles to a ServiceNow-ready Excel import file.

Reads from the "kb (YAML)" folder — articles use YAML frontmatter for metadata
and standard markdown for body content.

Sheet 1 – Articles      : one row per article, fields mapped to kb_knowledge columns.
                          Now self-tags each article with the Yale publish plan
                          (audience, per-instance availability, publish_plan, phase,
                          availability note) via kb_yale_classification.classify().
                          workflow_state is DERIVED from the plan, not hardcoded.
Sheet 2 – Relationships : one row per cross-reference (Prerequisites + Related Topics)
Sheet 3 – Import Instructions
"""

import re
import yaml
import markdown
from pathlib import Path
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from kb_yale_classification import classify   # Yale publish-plan tagging

# ── Instance configuration ────────────────────────────────────────────────────
# Paths are loaded from config.local.yaml in the repo root.
# Copy config.example.yaml → config.local.yaml and fill in your values.
# Falls back to sensible defaults relative to this script's location if the
# config file is not present (useful for running from within the repo directly).

_REPO_ROOT = Path(__file__).resolve().parent.parent
_CONFIG_PATH = _REPO_ROOT / "config.local.yaml"

def _load_config() -> dict:
    if _CONFIG_PATH.exists():
        with _CONFIG_PATH.open(encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    return {}

_cfg = _load_config()
_etl = _cfg.get("etl", {})

KB_DIR  = Path(_etl.get("kb_dir",  _REPO_ROOT / "kb (YAML)"))
OUT_FILE = Path(_etl.get("out_file", _REPO_ROOT / "ServiceNow ETL" / "REDCap_KB_ServiceNow_Import.xlsx"))

# Regex to extract RC-xxx IDs from strings like "RC-FD-02 — Online Designer"
REF_RE = re.compile(r"(RC-[A-Z0-9-]+)")


def parse_article(path: Path) -> dict:
    text = path.read_text(encoding="utf-8")

    # --- Split YAML frontmatter from markdown body ---
    front = {}
    body_text = text.strip()
    if text.startswith("---"):
        parts = text.split("---", 2)
        if len(parts) >= 3:
            front = yaml.safe_load(parts[1]) or {}
            body_text = parts[2].strip()

    # --- Core fields from frontmatter ---
    source_id  = str(front.get("id", path.stem.split("_")[0]))
    title      = str(front.get("title", ""))
    domain     = str(front.get("domain", ""))
    version    = str(front.get("version", ""))
    last_updated = str(front.get("last_updated", ""))

    # applies_to: list → semicolon-separated string
    applies_to_raw = front.get("applies_to", [])
    if isinstance(applies_to_raw, list):
        applies_to = "; ".join(str(x) for x in applies_to_raw)
    else:
        applies_to = str(applies_to_raw)

    # tags + synonyms → ServiceNow "meta" keyword field (search/index terms)
    tags_raw = front.get("tags", [])
    if not isinstance(tags_raw, list):
        tags_raw = [tags_raw] if tags_raw else []
    synonyms_raw = front.get("synonyms", [])
    if not isinstance(synonyms_raw, list):
        synonyms_raw = [synonyms_raw] if synonyms_raw else []
    meta = "; ".join(str(x) for x in (list(tags_raw) + list(synonyms_raw)) if str(x).strip())

    # --- Prerequisites ---
    prereqs_raw = front.get("prerequisites", [])
    if not isinstance(prereqs_raw, list):
        prereqs_raw = [prereqs_raw] if prereqs_raw else []
    # Filter out bare "None" entries
    prereqs_filtered = [p for p in prereqs_raw if str(p).strip().lower() != "none"]
    prerequisite_raw = "; ".join(str(p) for p in prereqs_raw)
    prereq_ids = []
    for p in prereqs_filtered:
        prereq_ids.extend(REF_RE.findall(str(p)))

    # --- Related topics ---
    related_raw = front.get("related", [])
    if not isinstance(related_raw, list):
        related_raw = [related_raw] if related_raw else []
    related_parts = []
    related_ids  = []
    for r in related_raw:
        if isinstance(r, dict):
            rid    = r.get("id", "")
            rtitle = r.get("title", "")
            related_parts.append(f"{rid} — {rtitle}" if rid and rtitle else rid or rtitle)
            if rid:
                related_ids.append(rid)
        else:
            s = str(r)
            related_parts.append(s)
            related_ids.extend(REF_RE.findall(s))
    related_topics_raw = "; ".join(related_parts)

    # --- Convert markdown body to HTML ---
    html_body = markdown.markdown(
        body_text,
        extensions=["tables", "fenced_code", "nl2br"]
    )

    # --- Yale publish classification (audience / per-instance / plan / phase) ---
    plan = classify(source_id)

    return {
        "source_id":          source_id,
        "short_description":  title,
        "category":           domain,
        # Yale publish-plan columns (self-tagged) ────────────────────────────
        "audience":           plan["audience"],
        "publish_plan":       plan["publish_plan"],
        "yale_phase":         plan["yale_phase"],
        "feature_dependent":  plan["feature_dependent"],
        "avail_rc1":          plan["avail_rc1"],
        "avail_rc2":          plan["avail_rc2"],
        "avail_connect":      plan["avail_connect"],
        "avail_p11":          plan["avail_p11"],
        "availability_note":  plan["availability_note"],
        # ─────────────────────────────────────────────────────────────────────
        "applies_to":         applies_to,
        "meta":               meta,        # tags + synonyms → kb_knowledge.meta
        "version":            version,
        "last_updated":       last_updated,
        "author":             "",          # not present in YAML frontmatter
        "prerequisite_raw":   prerequisite_raw,
        "related_topics_raw": related_topics_raw,
        "workflow_state":     plan["workflow_state"],   # DERIVED from publish_plan
        "kb_knowledge_base":  "",          # to be filled by ServiceNow admin
        "text":               html_body,
        "_prereq_ids":        prereq_ids,
        "_related_ids":       related_ids,
        "_synonyms":          [str(s) for s in synonyms_raw if str(s).strip()],
        "_path":              str(path.name),
    }


def style_header_row(ws, row=1, color="1F4E79"):
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=color)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def add_border(ws):
    thin = Side(style="thin", color="CCCCCC")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# Fills for quick visual triage on the Articles sheet
_PLAN_FILL = {
    "Publish":        "C6EFCE",
    "Draft (banner)": "FFEB9C",
    "Draft (verify)": "FFEB9C",
    "Hold":           "F8CBAD",
    "Hold (admin KB)":"FCE4D6",
    "Exclude":        "D9D9D9",
}
_AVAIL_FILL = {
    "Off":    "F8CBAD",
    "On*":    "FFF2CC",
    "Verify": "FFF2CC",
    "Part":   "FFF2CC",
}


def build_excel(articles: list[dict]):
    wb = Workbook()

    # ── Sheet 1: Articles ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Articles"

    art_headers = [
        "source_id",          # your RC-xxx ID — store in u_source_id custom field
        "short_description",  # article title → kb_knowledge.short_description
        "category",           # domain → kb_knowledge.category
        "audience",           # End-user | Power-user | Admin | Exclude  (Yale plan)
        "publish_plan",       # Publish | Draft (banner) | Hold | Hold (admin KB) | Exclude
        "yale_phase",         # 0 pilot | 1 core | 2 gated | 3 admin | — held
        "feature_dependent",  # Yes/No — availability varies by instance
        "avail_rc1",          # RC1 = redcap.med.yale.edu
        "avail_rc2",          # RC2 = redcap.research.yale.edu
        "avail_connect",      # Connect = redcapynh.ynhh.org
        "avail_p11",          # P11 = redcapynh-p11.ynhh.org (21 CFR Part 11)
        "availability_note",  # action / caveat for this article
        "kb_knowledge_base",  # leave blank; fill in after admin configures SN
        "workflow_state",     # DERIVED from publish_plan (published/draft/retired)
        "author",
        "version",
        "last_updated",
        "applies_to",         # maps to kb_knowledge.meta_description or custom field
        "meta",               # tags + synonyms → kb_knowledge.meta (search keywords)
        "prerequisite_raw",   # human-readable; for reference
        "related_topics_raw", # human-readable; for reference
        "text",               # HTML body → kb_knowledge.text
    ]

    ws1.append(art_headers)
    style_header_row(ws1, row=1, color="1F4E79")
    ws1.row_dimensions[1].height = 30

    # Column index lookups (1-based)
    col_idx = {h: i + 1 for i, h in enumerate(art_headers)}
    text_col = col_idx["text"]

    for art in articles:
        ws1.append([art.get(h, "") for h in art_headers])
        r = ws1.max_row
        # plan colour
        fill = _PLAN_FILL.get(art.get("publish_plan"))
        if fill:
            ws1.cell(row=r, column=col_idx["publish_plan"]).fill = PatternFill("solid", start_color=fill)
        # per-instance colour
        for h in ("avail_rc1", "avail_rc2", "avail_connect", "avail_p11"):
            af = _AVAIL_FILL.get(art.get(h))
            if af:
                ws1.cell(row=r, column=col_idx[h]).fill = PatternFill("solid", start_color=af)

    # Column widths by header name
    width_by_header = {
        "source_id": 14, "short_description": 42, "category": 18, "audience": 15,
        "publish_plan": 15, "yale_phase": 7, "feature_dependent": 11,
        "avail_rc1": 11, "avail_rc2": 12, "avail_connect": 13, "avail_p11": 12,
        "availability_note": 52, "kb_knowledge_base": 20, "workflow_state": 14,
        "author": 16, "version": 9, "last_updated": 13, "applies_to": 30,
        "meta": 34, "prerequisite_raw": 34, "related_topics_raw": 34, "text": 80,
    }
    for h, w in width_by_header.items():
        ws1.column_dimensions[get_column_letter(col_idx[h])].width = w

    # Wrap text and top-align data rows (skip wrap on the HTML text column)
    for row in ws1.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column != text_col))
            if cell.font is None or cell.font.name != "Arial":
                cell.font = Font(name="Arial", size=9)
        ws1.row_dimensions[row[0].row].height = 40

    ws1.freeze_panes = "D2"       # keep ID + title + category visible while scrolling
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(art_headers))}{ws1.max_row}"

    # ── Sheet 2: Relationships ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Relationships")

    rel_headers = [
        "source_id",           # article that contains the reference
        "source_title",
        "relationship_type",   # Prerequisite | Related Topic
        "target_id",           # RC-xxx being referenced
        "target_title",        # resolved if that article is also in the set
        "target_publish_plan", # Yale plan for the target — flags links to held/draft targets
        "notes",
    ]

    ws2.append(rel_headers)
    style_header_row(ws2, row=1, color="375623")
    ws2.row_dimensions[1].height = 22

    # Build lookups: source_id → title, source_id → publish_plan
    title_map = {a["source_id"]: a["short_description"] for a in articles}
    plan_map  = {a["source_id"]: a["publish_plan"] for a in articles}

    def _target_plan(tid):
        # match on short id (frontmatter ids may be short e.g. RC-FD-02)
        return plan_map.get(tid, classify(tid)["publish_plan"])

    for art in articles:
        sid = art["source_id"]
        stitle = art["short_description"]
        for tid in art["_prereq_ids"]:
            tplan = _target_plan(tid)
            warn = "⚠ target not Published — link would dead-end" if tplan != "Publish" else ""
            ws2.append([
                sid, stitle, "Prerequisite", tid,
                title_map.get(tid, "⚠ Not in current export"),
                tplan,
                (warn or "Must be read before this article")
            ])
        for tid in art["_related_ids"]:
            tplan = _target_plan(tid)
            warn = "⚠ target not Published — link would dead-end" if tplan != "Publish" else ""
            ws2.append([
                sid, stitle, "Related Topic", tid,
                title_map.get(tid, "⚠ Not in current export"),
                tplan,
                warn
            ])

    rel_widths = {"A": 16, "B": 42, "C": 18, "D": 16, "E": 42, "F": 16, "G": 40}
    for col, w in rel_widths.items():
        ws2.column_dimensions[col].width = w

    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Arial", size=10)
        ws2.row_dimensions[row[0].row].height = 30

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:G{ws2.max_row}"

    # ── Sheet 3: Instructions ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Import Instructions")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 90

    instructions = [
        ("REDCap KB → ServiceNow Import Guide", ""),
        ("", ""),
        ("SHEET: Articles", ""),
        ("source_id", "Store in a custom field (u_source_id) so articles remain findable by RC-xxx ID after import."),
        ("short_description", "Maps to kb_knowledge.short_description (the article title)."),
        ("category", "Maps to kb_knowledge.category. Your SN admin may need to create matching category records first."),
        ("audience", "Yale plan: End-user articles go in the end-user KB now; Admin/Power-user articles are held for a later admin KB; Exclude = not applicable to Yale."),
        ("publish_plan", "Yale rollout decision. Publish = load & publish. Draft (banner) = load hidden, add a Yale availability note, publish when the feature is confirmed. Hold / Hold (admin KB) = not for the end-user KB yet. Exclude = do not load."),
        ("yale_phase", "Rollout order: 0 pilot, 1 core end-user, 2 feature-gated/advanced, 3 admin KB, — held/excluded."),
        ("avail_rc1..p11", "Per-instance feature availability from config-scan section 6. Standard=always on; On=enabled; On*=enabled but per-project admin activation; Off=disabled; Verify/Part=confirm. Drives the availability banner text on feature articles."),
        ("availability_note", "Article-specific action or caveat (e.g. 'disabled on P11', 'Connect-only', 'populate from config scan')."),
        ("kb_knowledge_base", "Fill in the sys_id or name of your target Knowledge Base before importing."),
        ("workflow_state", "DERIVED from publish_plan: Publish→published, Draft*→draft, Hold*→draft, Exclude→retired. Override in SN if your review process differs."),
        ("author", "Must match a valid ServiceNow user (by name or sys_id). Update before import."),
        ("text", "HTML body. Paste into kb_knowledge.text. Verify formatting in SN preview after import."),
        ("", ""),
        ("SHEET: Relationships", ""),
        ("target_publish_plan", "The Yale plan for the linked target. Any value other than 'Publish' means the target is not (yet) live — wiring that link now would dead-end. Wire it up when the target is published."),
        ("How to use", "After Pass 1 import, use this sheet to wire up Related Articles links in ServiceNow — skipping/deferring links whose target_publish_plan is not 'Publish'."),
        ("⚠ Not in current export", "Target article was referenced but its source .md file was not in this export batch. Import it separately."),
        ("Two-pass approach", "1) Import Articles sheet → note the KB article numbers SN assigns.\n2) Use Relationships sheet + those numbers to create related-article links via SN admin or script."),
        ("", ""),
        ("RECOMMENDED CUSTOM FIELD", ""),
        ("u_source_id", "Add this custom field to kb_knowledge in ServiceNow. Store the RC-xxx ID here. Enables reliable lookups and future re-imports without duplication."),
        ("", ""),
        ("PUBLISH PLAN SOURCE", ""),
        ("kb_yale_classification.py", "The audience / per-instance / publish_plan / phase values are generated by kb_yale_classification.py, sourced from repo/current-projects/config-scan/instance-settings-checklist.md. Update that module when Yale enablement changes, then re-run this export."),
    ]

    ws3.append(["Field / Topic", "Guidance"])
    style_header_row(ws3, row=1, color="4A4A4A")
    ws3.row_dimensions[1].height = 22

    for i, (field, guidance) in enumerate(instructions, start=2):
        ws3.cell(row=i, column=1, value=field)
        ws3.cell(row=i, column=2, value=guidance)
        if field in ("SHEET: Articles", "SHEET: Relationships", "RECOMMENDED CUSTOM FIELD", "PUBLISH PLAN SOURCE"):
            for col in [1, 2]:
                ws3.cell(row=i, column=col).font = Font(name="Arial", size=10, bold=True)
                ws3.cell(row=i, column=col).fill = PatternFill("solid", start_color="D9E1F2")
        else:
            for col in [1, 2]:
                ws3.cell(row=i, column=col).font = Font(name="Arial", size=10)
        ws3.row_dimensions[i].height = 36

    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(OUT_FILE)
    print(f"Saved: {OUT_FILE}")
    return OUT_FILE


def _normalize_synonym(s: str) -> str:
    """Lowercase, collapse whitespace, strip surrounding punctuation for comparison."""
    s = str(s).lower().strip()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"^[^a-z0-9]+|[^a-z0-9]+$", "", s)
    return s


def check_synonym_collisions(articles: list[dict]) -> dict:
    """Build an inverted index of normalized synonym -> [article ids] and report
    any synonym claimed by more than one article. Returns the collision map."""
    index: dict[str, list] = {}
    for a in articles:
        sid = a["source_id"]
        seen_here = set()
        for raw in a.get("_synonyms", []):
            norm = _normalize_synonym(raw)
            if not norm or norm in seen_here:
                continue
            seen_here.add(norm)
            index.setdefault(norm, []).append((sid, raw))

    collisions = {k: v for k, v in index.items() if len(v) > 1}

    print("\n" + "=" * 70)
    print("SYNONYM COLLISION REPORT")
    print("=" * 70)
    total_syn = sum(len(a.get("_synonyms", [])) for a in articles)
    print(f"Articles with synonyms : {sum(1 for a in articles if a.get('_synonyms'))}")
    print(f"Total synonyms         : {total_syn}")
    print(f"Unique (normalized)    : {len(index)}")
    print(f"Colliding phrases      : {len(collisions)}")

    if not collisions:
        print("\nNo collisions found — every synonym maps to exactly one article.")
    else:
        print("\nThe following phrases are claimed by more than one article.")
        print("Resolve by: (a) qualifying each to its own variant, (b) assigning")
        print("one canonical owner, or (c) keeping if both are genuinely relevant.\n")
        for norm in sorted(collisions):
            owners = collisions[norm]
            print(f'  "{norm}"')
            for sid, raw in owners:
                print(f"      - {sid}: \"{raw}\"")
            print("")
    print("=" * 70 + "\n")
    return collisions


def check_title_collisions(articles: list[dict]) -> dict:
    """short_description (title) is the import coalesce/join key now that u_source_id
    is off the table. It MUST be unique. Report any duplicate titles."""
    index: dict[str, list] = {}
    for a in articles:
        t = re.sub(r"\s+", " ", str(a["short_description"]).strip().lower())
        index.setdefault(t, []).append(a["source_id"])
    collisions = {k: v for k, v in index.items() if len(v) > 1}
    print("=" * 70)
    print("TITLE (short_description) UNIQUENESS CHECK")
    print("=" * 70)
    print("short_description is the ServiceNow coalesce key — it must be unique.")
    if not collisions:
        print("OK — every article title is unique.\n")
    else:
        print(f"⚠ {len(collisions)} duplicate title(s) — these WILL collide on import:\n")
        for t, ids in sorted(collisions.items()):
            print(f'  "{t}"  ->  {", ".join(ids)}')
        print("")
    print("=" * 70 + "\n")
    return collisions


def report_publish_plan(articles: list[dict]) -> None:
    """Print the Yale publish-plan summary so a re-run shows the current split."""
    plan_c  = Counter(a["publish_plan"] for a in articles)
    phase_c = Counter(a["yale_phase"] for a in articles)
    aud_c   = Counter(a["audience"] for a in articles)
    print("=" * 70)
    print("YALE PUBLISH PLAN SUMMARY")
    print("=" * 70)
    print(f"Total articles: {len(articles)}")
    print("\nBy publish_plan:")
    for k in ["Publish","Draft (banner)","Draft (verify)","Hold","Hold (admin KB)","Exclude"]:
        if plan_c.get(k):
            print(f"  {k:18s} {plan_c[k]:>4}")
    print("\nBy phase:")
    for k in ["0","1","2","3","—"]:
        if phase_c.get(k):
            label = {"0":"0 pilot","1":"1 core end-user","2":"2 feature-gated/advanced",
                     "3":"3 admin KB","—":"held/excluded"}[k]
            print(f"  {label:26s} {phase_c[k]:>4}")
    print("\nBy audience:")
    for k, v in aud_c.most_common():
        print(f"  {k:18s} {v:>4}")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    md_files = sorted(KB_DIR.glob("RC-*.md"))
    print(f"Found {len(md_files)} KB articles:")
    for f in md_files:
        print(f"  {f.name}")

    articles = [parse_article(f) for f in md_files]

    print("\nArticle summary:")
    for a in articles:
        print(f"  {a['source_id']:12s} | {a['audience']:16s} | {a['publish_plan']:15s} "
              f"| ph{a['yale_phase']} | prereqs={a['_prereq_ids']} | related={a['_related_ids']}")

    report_publish_plan(articles)
    check_title_collisions(articles)
    check_synonym_collisions(articles)

    build_excel(articles)
    print("Done.")
