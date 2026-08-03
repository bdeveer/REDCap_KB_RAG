"""
resolve_links.py  —  Pass 2 of the REDCap KB → ServiceNow load.

Run this AFTER the Articles sheet has been imported into ServiceNow and the SN
admin has sent back an export of the imported articles. It turns the RC-xxx
references inside each article into real ServiceNow hyperlinks, and produces a
ready-to-import "Related Articles" file for the sidebar.

Because Yale is NOT adding a u_source_id custom field, the join key across the
two passes is the article TITLE (short_description). The handback file maps each
title to its ServiceNow sys_id; every link is built from that.

────────────────────────────────────────────────────────────────────────────────
INPUTS
  1. The KB articles (same "kb (YAML)" folder the export reads).
  2. A handback CSV/XLSX exported from kb_knowledge after Pass 1, with at least:
        - a title column     (short_description / title / name)
        - a sys_id column     (sys_id / sys_kb_id)
     A "number" column (KB00…) is used if present, otherwise ignored.

OUTPUTS  (written next to this script)
  - Articles_LinkUpdate.xlsx : short_description + rewritten HTML body.
        Re-import this coalescing on short_description to update bodies in place.
  - Related_Articles_Import.xlsx : source/target sys_ids for the native
        Related Articles sidebar (published→published pairs only).

LINK POLICY
  - Only link to targets whose Yale publish_plan == "Publish" (an end user should
    not be sent to a draft/held article). References to not-yet-published targets
    are left as plain text and listed in the report so a later pass can fill them
    in once those articles go live.
────────────────────────────────────────────────────────────────────────────────
"""

import re
import csv
from pathlib import Path
from collections import Counter
from openpyxl import Workbook, load_workbook

from kb_to_servicenow import parse_article, KB_DIR, style_header_row
from kb_yale_classification import classify

# Confirmed Yale Service Portal article URL pattern (keyed on sys_id):
URL_TEMPLATE = "https://yale.service-now.com/kb?id=kb_article_view&sys_kb_id={sys_id}"

HANDBACK_FILE = Path(__file__).resolve().parent / "sn_handback.csv"   # override as needed
OUT_DIR = Path(__file__).resolve().parent

# Matches RC-FD-02, RC-API-56, RC-AT-EM-01, RC-NAV-REC-04, RC-CDIS-03, …
RCID_RE = re.compile(r"RC-(?:[A-Z]+-)+\d+")
# Splits HTML into anchor tags vs. everything else, so we never rewrite inside a link
ANCHOR_SPLIT_RE = re.compile(r"(<a\b[^>]*>.*?</a>)", re.IGNORECASE | re.DOTALL)


def _norm_title(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())


def load_handback(path: Path) -> dict:
    """Return {normalized_title: sys_id} from the admin's post-import export."""
    if not path.exists():
        raise FileNotFoundError(
            f"Handback file not found: {path}\n"
            "Provide the SN export (title + sys_id) from after Pass 1."
        )
    rows = []
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(header, r)))
    else:
        with path.open(encoding="utf-8-sig", newline="") as f:
            for r in csv.DictReader(f):
                rows.append({(k or "").strip().lower(): v for k, v in r.items()})

    def pick(d, *names):
        for n in names:
            for k, v in d.items():
                if k == n:
                    return v
        return None

    mapping = {}
    for d in rows:
        title = pick(d, "short_description", "title", "name")
        sysid = pick(d, "sys_id", "sys_kb_id")
        if title and sysid:
            mapping[_norm_title(title)] = str(sysid).strip()
    return mapping


def rewrite_body(html: str, rcid_to_url: dict, self_id: str) -> tuple[str, list]:
    """Wrap resolvable RC-xxx tokens in <a href>. Returns (new_html, unresolved_ids)."""
    unresolved = []

    def repl(m):
        rid = m.group(0)
        if rid == self_id:
            return rid
        url = rcid_to_url.get(rid)
        if url:
            return f'<a href="{url}">{rid}</a>'
        unresolved.append(rid)
        return rid

    out_parts = []
    for seg in ANCHOR_SPLIT_RE.split(html):
        if seg[:2].lower() == "<a":          # existing anchor — leave untouched
            out_parts.append(seg)
        else:
            out_parts.append(RCID_RE.sub(repl, seg))
    return "".join(out_parts), unresolved


def main():
    articles = [parse_article(f) for f in sorted(KB_DIR.glob("RC-*.md"))]
    by_id = {a["source_id"]: a for a in articles}

    handback = load_handback(HANDBACK_FILE)

    # RC-id → sys_id, joining our title to the handback title
    rcid_to_sysid, missing_from_handback = {}, []
    for a in articles:
        sysid = handback.get(_norm_title(a["short_description"]))
        if sysid:
            rcid_to_sysid[a["source_id"]] = sysid
        else:
            missing_from_handback.append(a["source_id"])

    # Only build URLs for targets that are meant to be PUBLISHED
    rcid_to_url = {
        rid: URL_TEMPLATE.format(sys_id=sid)
        for rid, sid in rcid_to_sysid.items()
        if classify(rid)["publish_plan"] == "Publish"
    }

    # ── Rewrite bodies (only for published source articles) ─────────────────
    updated, all_unresolved = [], Counter()
    for a in articles:
        if a["publish_plan"] != "Publish":
            continue
        new_html, unresolved = rewrite_body(a["text"], rcid_to_url, a["source_id"])
        for u in unresolved:
            all_unresolved[u] += 1
        if new_html != a["text"]:
            updated.append((a["short_description"], new_html))

    wb1 = Workbook(); ws1 = wb1.active; ws1.title = "Articles_LinkUpdate"
    ws1.append(["short_description", "text"])
    style_header_row(ws1, 1, "1F4E79")
    for title, html in updated:
        ws1.append([title, html])
    ws1.column_dimensions["A"].width = 45
    ws1.column_dimensions["B"].width = 100
    out1 = OUT_DIR / "Articles_LinkUpdate.xlsx"
    wb1.save(out1)

    # ── Related Articles sidebar (published → published only) ────────────────
    wb2 = Workbook(); ws2 = wb2.active; ws2.title = "Related_Articles"
    ws2.append(["source_short_description", "source_sys_id",
                "target_short_description", "target_sys_id",
                "relationship_type", "note"])
    style_header_row(ws2, 1, "375623")
    rel_count, rel_skipped = 0, 0
    for a in articles:
        if a["publish_plan"] != "Publish":
            continue
        s_sys = rcid_to_sysid.get(a["source_id"])
        pairs = [(t, "Prerequisite") for t in a["_prereq_ids"]] + \
                [(t, "Related Topic") for t in a["_related_ids"]]
        for tid, rtype in pairs:
            t_art = by_id.get(tid)
            t_pub = (t_art["publish_plan"] if t_art else classify(tid)["publish_plan"])
            t_sys = rcid_to_sysid.get(tid)
            if t_pub == "Publish" and s_sys and t_sys:
                ws2.append([a["short_description"], s_sys,
                            (t_art["short_description"] if t_art else tid), t_sys,
                            rtype, ""])
                rel_count += 1
            else:
                rel_skipped += 1
    for col, w in {"A":42,"B":34,"C":42,"D":34,"E":16,"F":30}.items():
        ws2.column_dimensions[col].width = w
    out2 = OUT_DIR / "Related_Articles_Import.xlsx"
    wb2.save(out2)

    # ── Report ───────────────────────────────────────────────────────────────
    print("=" * 70)
    print("LINK RESOLUTION REPORT")
    print("=" * 70)
    print(f"Articles parsed              : {len(articles)}")
    print(f"Matched to handback (sys_id) : {len(rcid_to_sysid)}")
    print(f"Publishable link targets     : {len(rcid_to_url)}")
    print(f"Article bodies rewritten     : {len(updated)}   -> {out1.name}")
    print(f"Related-article links written: {rel_count}   -> {out2.name}")
    print(f"Related links skipped        : {rel_skipped} (target not published or no sys_id)")
    if missing_from_handback:
        print(f"\n⚠ {len(missing_from_handback)} article(s) had no title match in the handback "
              f"(check for renamed titles):")
        for i in missing_from_handback[:20]:
            print(f"    {i}  — \"{by_id[i]['short_description']}\"")
        if len(missing_from_handback) > 20:
            print(f"    … and {len(missing_from_handback) - 20} more")
    if all_unresolved:
        print(f"\nℹ In-body references left as plain text (target not yet published), top 15:")
        for rid, n in all_unresolved.most_common(15):
            print(f"    {rid}: {n}  ({classify(rid)['publish_plan']})")
    print("=" * 70)
    print("Next: re-import Articles_LinkUpdate.xlsx coalescing on short_description "
          "(updates bodies in place); import Related_Articles_Import.xlsx into the "
          "related-articles table using the sys_id columns.")


if __name__ == "__main__":
    main()
